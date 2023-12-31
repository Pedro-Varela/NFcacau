import xml.etree.ElementTree as ET
import openpyxl
import os

def registrar_numero_nota(numero_nota, arquivo_registro='notas_processadas.txt'):
    with open(arquivo_registro, 'a') as file:
        file.write(numero_nota + '\n')

def verificar_numero_nota(numero_nota, arquivo_registro='notas_processadas.txt'):
    try:
        with open(arquivo_registro, 'r') as file:
            notas_processadas = file.read().splitlines()
            return numero_nota in notas_processadas
    except FileNotFoundError:
        return False

def adicionar_dados_ao_excel_produtos(arquivo_excel, numero_nota, dados_produtos):
    # Verificar se o arquivo Excel existe
    if os.path.exists(arquivo_excel):
        workbook = openpyxl.load_workbook(arquivo_excel)
    else:
        workbook = openpyxl.Workbook()
        # Adicionar cabeçalhos se for um novo arquivo
        headers = ['NúmeroNota',
                   'EAN',
                   'NCM',
                   'CEST', 
                   'Código', 
                   'Descrição',
                   'QTD',
                   'Vunit', 
                   'VTotal',
                   'nLote', 
                   'dFab', 
                   'dVal',
                   'vICMS',
                   'pICMS', 
                   'vIPI', 
                   'pIPI',
                   'vPIS', 
                   'pPIS', 
                   'vCOFINS', 
                   'pCOFINS']
        
        workbook.active.append(headers)

    sheet = workbook.active

    # Adicionar os dados dos produtos
    for produto in dados_produtos:
        row = [
            numero_nota,
            produto['cEAN'],
            produto['NCM'],
            produto['CEST'],
            produto['cProd'], 
            produto['xProd'], 
            produto['qCom'],
            produto['vUnCom'], 
            produto['vProd'],   
            produto['nLote'],
            produto['dFab'],
            produto['dVal'],
            produto['vICMS'],
            produto['pICMS'],
            produto['vIPI'],
            produto['pIPI'],
            produto['vPIS'],
            produto['pPIS'],
            produto['vCOFINS'],
            produto['pCOFINS']
        ]
        sheet.append(row)


    workbook.save(arquivo_excel)
    # Salvar o arquivo Excel
        


def adicionar_dados_ao_excel_notas(arquivo_excel, numero_nota, dados_notas):
    # Verificar se o arquivo Excel existe
    if os.path.exists(arquivo_excel):
        workbook = openpyxl.load_workbook(arquivo_excel)
    else:
        workbook = openpyxl.Workbook()
        # Adicionar cabeçalhos se for um novo arquivo
        headers = [
            'Número da Nota', 'Data de Emissão', 'Data de Saída/Entrada', 'Código Município', 'Tipo de Impressão', 
            'CNPJ Emitente', 'Nome Emitente', 'Nome Fantasia Emitente', 'IE Emitente', 'Município Emitente', 'UF Emitente', 'CNAE Emitente', 
            'Valor Total da Nota', 'Valor Total dos Produtos', 'Valor Total do ICMS', 'Valor Total do PIS', 'Valor Total do COFINS', 
            'Nome Transportador', 'CNPJ Transportador'
        ]
        workbook.active.append(headers)

    sheet = workbook.active

    # Adicionar os dados das notas
    for nota in dados_notas:
        row = [
            numero_nota, 
            nota['dhEmi'],  
            nota['dhSaiEnt'],  
            nota['cMunFG'],  
            nota['tpImp'],  
            nota['CNPJ_emit'],  
            nota['xNome_emit'],  
            nota['xFant_emit'],  
            nota['IE_emit'],  
            nota['xMun_emit'],  
            nota['UF_emit'],  
            nota['CNAE_emit'],  
            nota['vNF'],   
            nota['vProd_totais'],  
            nota['vICMS_totais'],    
            nota['vPIS_totais'],    
            nota['vCOFINS_totais'],    
            nota['xNome_transp'],  
            nota['CNPJ_transp']  
        ]
        sheet.append(row)

    # Salvar o arquivo Excel
    workbook.save(arquivo_excel)



