import xml.etree.ElementTree as ET
import openpyxl
import os

def registrar_numero_nota(numero_nota, arquivo_registro='notas_processadas.txt'):
    with open(arquivo_registro, 'a') as file:
        file.write(numero_nota + '\n')

def verificar_numero_nota(numero_nota, arquivo_registro='notas_processadas.txt'):
    try:
        with open(arquivo_registro, 'r') as file:
            notas_processadas = file.read().splitlines()
            return numero_nota in notas_processadas
    except FileNotFoundError:
        return False

def adicionar_dados_ao_excel_produtos(arquivo_excel, numero_nota, dados_produtos):
    # Verificar se o arquivo Excel existe
    if os.path.exists(arquivo_excel):
        workbook = openpyxl.load_workbook(arquivo_excel)
    else:
        workbook = openpyxl.Workbook()
        # Adicionar cabeçalhos se for um novo arquivo
        headers = ['NúmeroNota',
                   'EAN',
                   'NCM',
                   'CEST', 
                   'Código', 
                   'Descrição',
                   'QTD',
                   'Unidade',
                   'Vunit', 
                   'VTotal',
                   'nLote', 
                   'dFab', 
                   'dVal',
                   'vICMS',
                   'pICMS', 
                   'vIPI', 
                   'pIPI',
                   'vPIS', 
                   'pPIS', 
                   'vCOFINS', 
                   'pCOFINS']
        
        workbook.active.append(headers)

    sheet = workbook.active

    # Adicionar os dados dos produtos
    for produto in dados_produtos:
        row = [
            numero_nota,
            produto['cEAN'],
            produto['NCM'],
            produto['CEST'],
            produto['cProd'], 
            produto['xProd'], 
            produto['qCom'],
            produto['uCom'],
            produto['vUnCom'], 
            produto['vProd'],   
            produto['nLote'],
            produto['dFab'],
            produto['dVal'],
            produto['vICMS'],
            produto['pICMS'],
            produto['vIPI'],
            produto['pIPI'],
            produto['vPIS'],
            produto['pPIS'],
            produto['vCOFINS'],
            produto['pCOFINS']
        ]
        sheet.append(row)


    workbook.save(arquivo_excel)
    # Salvar o arquivo Excel
        


def adicionar_dados_ao_excel_notas(arquivo_excel, numero_nota, dados_notas):
    # Verificar se o arquivo Excel existe
    if os.path.exists(arquivo_excel):
        workbook = openpyxl.load_workbook(arquivo_excel)
    else:
        workbook = openpyxl.Workbook()
        # Adicionar cabeçalhos se for um novo arquivo
        headers = [
            'Número da Nota', 'Data de Emissão', 'Data de Saída/Entrada','DataVenc1','DataVenc2', 'Código Município', 'Tipo de Impressão', 
            'CNPJ Emitente', 'Nome Emitente', 'Nome Fantasia Emitente', 'IE Emitente', 'Município Emitente', 'UF Emitente', 'CNAE Emitente', 
            'Valor Total da Nota', 'Valor Total dos Produtos', 'Valor Total do ICMS', 'Valor Total do PIS', 'Valor Total do COFINS', 
            'Nome Transportador', 'CNPJ Transportador'
        ]
        workbook.active.append(headers)

    sheet = workbook.active


    # Adicionar os dados das notas
    for nota in dados_notas:
        datas_vencimento = nota.get('datasVencimento', [])
        data_venc_1 = datas_vencimento[0] if len(datas_vencimento) > 0 else ''
        data_venc_2 = datas_vencimento[1] if len(datas_vencimento) > 1 else ''
        
        row = [
          numero_nota,
            nota.get('dhEmi', ''),  
            nota.get('dhSaiEnt', ''),  
            data_venc_1,
            data_venc_2,
            nota.get('cMunFG', ''),  
            nota.get('tpImp', ''),  
            nota.get('CNPJ_emit', ''),  
            nota.get('xNome_emit', ''),  
            nota.get('xFant_emit', ''),  
            nota.get('IE_emit', ''),  
            nota.get('xMun_emit', ''),  
            nota.get('UF_emit', ''),  
            nota.get('CNAE_emit', ''),  
            nota.get('vNF', ''),   
            nota.get('vProd_totais', ''),  
            nota.get('vICMS_totais', ''),    
            nota.get('vPIS_totais', ''),    
            nota.get('vCOFINS_totais', ''),    
            nota.get('xNome_transp', ''),  
            nota.get('CNPJ_transp', '')   
        ]
        sheet.append(row)

    # Salvar o arquivo Excel
    workbook.save(arquivo_excel)



